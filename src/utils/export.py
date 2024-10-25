
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from ..config import COLORS, MANUFACTURERS

class ExcelExporter:
    def __init__(self, project):
        self.project = project
        self.setup_styles()

    def setup_styles(self):
        # Stile bordo
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Stile header
        self.header_fill = PatternFill(
            start_color=COLORS['primary'].replace('#', ''), 
            end_color=COLORS['primary'].replace('#', ''), 
            fill_type="solid"
        )
        self.header_font = Font(color="FFFFFF", bold=True)

        # Stile numeri
        self.number_style = NamedStyle(
            name='number_style',
            number_format='#,##0.00',
            alignment=Alignment(horizontal='right')
        )

        # Stile testo
        self.text_style = NamedStyle(
            name='text_style',
            alignment=Alignment(horizontal='left')
        )

    def export(self, filename):
        try:
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                self._write_info_sheet(writer)
                
                for tableau_name, tableau in self.project.tableaux.items():
                    # Scrive un foglio per ogni produttore
                    for manufacturer in MANUFACTURERS:
                        self._write_manufacturer_sheet(writer, tableau_name, tableau, manufacturer)
                    
                    self._write_labor_sheet(writer, tableau_name, tableau)
                    self._write_summary_sheet(writer, tableau_name, tableau)
                
                self._apply_styles(writer)
                self._adjust_columns(writer)
            return True
        except Exception as e:
            print(f"Errore durante l'esportazione: {str(e)}")
            return False

    def _write_info_sheet(self, writer):
        info_data = {
            'Progetto': [self.project.name],
            'Cliente': [self.project.client_name],
            'Numero Volta': [self.project.volta_number],
            'Data Creazione': [self.project.creation_date]
        }
        df = pd.DataFrame(info_data)
        df.to_excel(writer, sheet_name='Info Progetto', index=False)

        # Aggiungi stili personalizzati
        sheet = writer.sheets['Info Progetto']
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
            for cell in row:
                cell.border = self.border
                if cell.row == 1:
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                cell.alignment = Alignment(horizontal='left')

    def _write_manufacturer_sheet(self, writer, tableau_name, tableau, manufacturer):
        if manufacturer not in tableau.materials_data or not tableau.materials_data[manufacturer]:
            return

        data = tableau.materials_data[manufacturer]
        df = pd.DataFrame(data)
        sheet_name = f'{tableau_name} - {manufacturer}'
        df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Formattazione colonne
        sheet = writer.sheets[sheet_name]
        for idx, column in enumerate(df.columns, 1):
            col_letter = get_column_letter(idx)
            # Calcola la larghezza massima
            max_length = max(
                max((len(str(value)) for value in df[column]), default=0),
                len(column)
            )
            sheet.column_dimensions[col_letter].width = min(max_length + 2, 50)

    def _write_labor_sheet(self, writer, tableau_name, tableau):
        if not tableau.labor_data:
            return
        
        df = pd.DataFrame(tableau.labor_data)
        sheet_name = f'{tableau_name} - Manodopera'
        df.to_excel(writer, sheet_name=sheet_name, index=False)

        sheet = writer.sheets[sheet_name]
        # Imposta larghezze colonne specifiche per manodopera
        sheet.column_dimensions['A'].width = 30  # Type
        sheet.column_dimensions['B'].width = 15  # Heures
        sheet.column_dimensions['C'].width = 15  # Tarif

    def _write_summary_sheet(self, writer, tableau_name, tableau):
        if not tableau.summary_data:
            return
        
        df = pd.DataFrame(tableau.summary_data)
        sheet_name = f'{tableau_name} - Riepilogo'
        df.to_excel(writer, sheet_name=sheet_name, index=False)

        sheet = writer.sheets[sheet_name]
        # Imposta stili specifici per il riepilogo
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')

    def _apply_styles(self, writer):
        for sheet_name in writer.sheets:
            sheet = writer.sheets[sheet_name]
            
            # Applica stili all'header
            for cell in sheet[1]:
                cell.fill = self.header_fill
                cell.font = self.header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = self.border

            # Applica stili alle celle dati
            for row in sheet.iter_rows(min_row=2):
                for cell in row:
                    cell.border = self.border
                    if isinstance(cell.value, (int, float)):
                        cell.style = self.number_style
                    else:
                        cell.style = self.text_style

    def _adjust_columns(self, writer):
        for sheet_name in writer.sheets:
            sheet = writer.sheets[sheet_name]
            for column in sheet.columns:
                max_length = 0
                column = list(column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[get_column_letter(column[0].column)].width = min(adjusted_width, 50)
